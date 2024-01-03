from email import message
#from fileinput import filename
from tkinter import *
from tkinter import ttk
import time
import sqlite3
from tkinter import messagebox
from tkinter.filedialog import askopenfilename
from turtle import width
from DataManager import Data_Window
from urllib.parse import urlencode
import requests
from openpyxl import load_workbook
from PIL import ImageTk, Image

with sqlite3.connect("Base_de_Datos/DataBase.db") as db:
    cursor = db.cursor()

#SECCIONES
Cocteles = ['SELECCIONA']
Platillos = ['SELECCIONA']
Caldos = ['SELECCIONA']
Pescados = ['SELECCIONA']
Bebidas = ['SELECCIONA']
Postres = ['SELECCIONA']
Meseros = ['SELECCIONA']
Tickets = []

#+++++++++++++++++++++++++++++FUNCIONES++++++++++++++++++++++++++++++#

# def Display_Time():
#     now = time.strftime("Fecha: %d/%b/%Y, Hora: %H:%M:%S")
#     Reloj_Eti['text'] = now
#     window.after(1000,Display_Time)

def Generador_de_Secciones():
    del Cocteles[:], Platillos[:], Caldos[:],Pescados[:], Bebidas[:], Postres[:], Meseros[:]
    Secciones = [Cocteles, Platillos, Caldos, Pescados, Bebidas, Postres, Meseros]
    Tables = ['COCTELES', 'PLATILLOS', 'CALDOS', 'PESCADOS', 'BEBIDAS', 'POSTRES','EMPLEADOS']
    count = 0
    for i in Tables:
        cursor.execute("SELECT Nombre FROM " + i + " ")
        for x in cursor.fetchall():
            Secciones[count].append(x[0])
        count += 1

    return Cocteles, Bebidas, Platillos, Caldos, Postres, Pescados, Meseros

def CheckDescuento():
    Lectura = str(Env1Ent1['state'])
    if Lectura == 'disabled':
        Env1Ent1['state'] = 'normal'
        pass
    else:
        Env1Ent1.delete(0,END)
        Env1Ent1['state'] = 'disable'
        Descuento.set(0)
    pass

def CheckImporte():
    Lectura = str(Env1Ent2['state'])
    if Lectura == 'disabled':
        Env1Ent2['state'] ='normal'
        pass
    else:
        Env1Ent2.delete(0,END)
        Env1Ent2['state'] = 'disable'
        Importe.set(0)
        
def CheckLlevar():
    Lectura1 = str(Ent_Personas['state'])
    Lectura2 = str(Mesas_Ent['state'])
    if Lectura1 and Lectura2 == 'normal':
        Mesa.set(0)
        Personas.set(0)
        Ent_Personas['state'] = 'disabled'
        Mesas_Ent['state'] = 'disabled'
        MasMesBoton['state'] = 'disabled'
        MenMesBoton['state'] = 'disabled'
        MasPerBoton['state'] = 'disabled'
        MenPerBoton['state'] = 'disabled'
        pass
    else:
        Ent_Personas['state'] = 'normal'
        Mesas_Ent['state'] = 'normal'
        MasMesBoton['state'] = 'normal'
        MenMesBoton['state'] = 'normal'
        MasPerBoton['state'] = 'normal'
        MenPerBoton['state'] = 'normal'
        pass

def Activar_Peso(*a):
    Lectura = str(Pescados_Lista.get())
    if not Lectura == 'SELECCIONA':
        Peso_Pescado['state'] = 'normal'
        pass
    else:
        Peso_Pescado.delete(0,END)
        Peso_Pescado['state'] = 'disable'

def Otro_Producto():
    Lec1, Lec2 = str(Nombre_Otro['state']), str(Precio_Otro['state'])
    
    if Lec1 and Lec2 == 'disabled':
        Nombre_Otro['state'] = 'normal'
        Precio_Otro['state'] = 'normal'
        pass
    else:
        Nombre_Otro.delete(0,END)
        Precio_Otro.delete(0,END)
        Nombre_Otro['state'] = 'disabled'
        Precio_Otro['state'] = 'disabled'

def SisCantidades(args):
    x = int(Cantidad.get())
    if args == 1:
        x = x + 1
        Cantidad.set(x)
        pass
    else:
        if x > 1:
            x = x - 1
            Cantidad.set(x)
            pass

def SisPersonas(event):
    x = int(Personas.get())
    if event == 1:
        x = x + 1
        Personas.set(x)
        pass
    else:
        if x > 0:
            x = x - 1
            Personas.set(x)
        pass

def SisMesas(event):
    x = int(Mesa.get())
    if event == 1:
        x = x +1
        Mesa.set(x)
        pass
    else:
        if x > 1:
            x = x - 1
            Mesa.set(x)
        pass

def Agregar():
    Name_Cocteles = Cocteles_Lista.get()
    Name_Bebidas = Bebidas_Lista.get()
    Name_Platillos = Platillos_Lista.get()
    Name_Caldos = Caldos_Lista.get()
    Name_Postres = Postres_Lista.get()
    Name_Pescados = Pescados_Lista.get()
    cant = Cantidad.get()
    peso = Peso_Pescado.get()
    Lec1, Lec2 = str(Nombre_Otro['state']), str(Precio_Otro['state'])

    if not Name_Cocteles == 'SELECCIONA':
        cursor.execute("SELECT COCTELES.Precio FROM COCTELES WHERE COCTELES.Nombre = ?",[Name_Cocteles])
        a = int(cursor.fetchone()[0])
        if  cant > 0:
            Precio = cant * a
            Comanda.insert("",END,text = 0, values = (cant,Name_Cocteles,Precio))
            pass
        pass
    elif not Name_Bebidas == 'SELECCIONA':
        cursor.execute("SELECT BEBIDAS.Precio FROM BEBIDAS WHERE BEBIDAS.Nombre = ?",[Name_Bebidas])
        b = int(cursor.fetchone()[0])
        if cant > 0:
            Precio = cant * b
            Comanda.insert("",END,text = 0, values = (cant,Name_Bebidas,Precio))  
            pass
        pass
    elif not Name_Platillos == 'SELECCIONA':
        cursor.execute("SELECT PLATILLOS.Precio FROM PLATILLOS WHERE PLATILLOS.Nombre = ?",[Name_Platillos])
        c = int(cursor.fetchone()[0])
        if cant > 0:
            Precio = cant * c
            Comanda.insert("",END,text = 0, values = (cant,Name_Platillos,Precio))  
            pass
        pass
    elif not Name_Caldos == 'SELECCIONA':
        cursor.execute("SELECT CALDOS.Precio FROM CALDOS WHERE CALDOS.Nombre = ?",[Name_Caldos])
        d = int(cursor.fetchone()[0])
        if cant > 0:
            Precio = cant * d
            Comanda.insert("",END,text = 0, values = (cant,Name_Caldos,Precio))  
            pass
        pass
    elif not Name_Postres == 'SELECCIONA':
        cursor.execute("SELECT POSTRES.Precio FROM POSTRES WHERE POSTRES.Nombre = ?",[Name_Postres])
        e = int(cursor.fetchone()[0])
        if cant > 0:
            Precio = cant * e
            Comanda.insert("",END,text = 0, values = (cant,Name_Postres,Precio))  
            pass
        pass
    elif not Name_Pescados == 'SELECCIONA':
        cursor.execute("SELECT PESCADOS.Precio FROM PESCADOS WHERE PESCADOS.Nombre = ?",[Name_Pescados])
        f = int(cursor.fetchone()[0])
        if cant > 0:
            peso = int(peso)
            Precio = cant * f * peso/1000
            w = str(peso)
            Comanda.insert("",END,text = 0, values = (cant,Name_Pescados + " " + w + "gr",Precio))  
            pass
        pass
    elif Lec1 and Lec2 == 'normal':
        Other = str(Nombre_Otro.get())
        Other_Price = float(Precio_Otro.get())
        
        if Other != '' and cant > 0 and Other_Price > 0:
            Precio = cant * Other_Price
            Comanda.insert("",END, text = 0, values = (cant,Other,Precio))
        pass
    pass

    Connect_Tot()

    Cocteles_Lista.set('SELECCIONA')
    Bebidas_Lista.set('SELECCIONA')
    Platillos_Lista.set('SELECCIONA')
    Caldos_Lista.set('SELECCIONA')
    Postres_Lista.set('SELECCIONA')
    Pescados_Lista.set('SELECCIONA')
    Nombre_Otro.delete(0,END)
    Precio_Otro.delete(0,END)
    Cantidad.set(1)

def Borrar(args):
    if args == 1:
        item = Comanda.selection()
        Comanda.delete(item)
        pass
    else:
        Lectura = Comanda.get_children()
        if not str(Lectura[0]) == '':
            Mensaje = messagebox.askquestion(title = 'Alerta', message = '¿Seguro que deseas borrar el contenido?')
            if Mensaje == 'yes':
                Comanda.delete(*Lectura)
                pass
            pass
    Connect_Tot()

def ConnectExcel():
    Excel_name = askopenfilename(title="Ubicación del Excel", filetypes=[("Excel (.xlsx)","*.xlsx")])
    sheet_name = str(time.strftime("%d-%m-%Y"))
    try:
        wb = load_workbook(Excel_name)
        ws = wb.create_sheet(sheet_name)
        Fecha = time.strftime("%d/%b/%Y") 
        cursor.execute("SELECT * FROM REGISTRO WHERE FECHA LIKE '%"+Fecha+"%' AND COD_VENTA > 0")
        Encabezados = ("Mesa","Tipo","ID Empleado","Ganancia","Importe","Descuento","Num. Personas","Metodo Pago","Factura","Fecha")
        ws.append(Encabezados)
        for item in cursor.fetchall():
            ws.append(item[1:])

        wb.save(Excel_name)
        messagebox.showinfo(title='Excel', message='Se ha generado el registro en Excel')
    except:
        messagebox.showerror(title="ERROR EXCEL", message="NO PUDO GENERAR EL EXCEL")

def Connect_Tot(*args):
    Imp = Importe.get()
    Des = Descuento.get()

    items = Comanda.get_children()
    
    Total_cost = 0
    
    Precios = []
    for i in items:
        Elemento = Comanda.item(i)
        Precio = Elemento['values'][2]
        Precios.append(Precio)
    #SI MIDE MENOS DE 20 HAY QUE RELLENARLA CON ESPACIOS (POR RESOLVER)
    #SI MIDE MÁS DE 20, HAY QUE RECORTARLO (RESUELTO)
    for row in Precios:
        Total_cost += float(row)
    
    if Imp != "":
        Total_cost += Imp
    if Des != "":
        Total_cost -= Des

    Total.set(Total_cost)

def UpdateListas():
    Cocteles, Bebidas, Platillos, Caldos, Postres, Pescados, Meseros  = Generador_de_Secciones()
    Cocteles_Lista["values"] = Cocteles
    Bebidas_Lista["values"] = Bebidas
    Platillos_Lista['values'] = Platillos
    Caldos_Lista["values"] = Caldos
    Pescados_Lista["values"] = Pescados
    Postres_Lista["values"] = Postres
    Meseros_Lista["values"] = Meseros

def Clear_WorkSpace():
    Set_Strings = [Coctel,Bebida,Platillo,Caldo,Postre,Pescado,Mesero]
    Set_Ints = [Personas,Mesa,Total,Importe,Descuento]
    Entradas = [Pagacon_Entry,Env1Ent1,Env1Ent2]
    for i in Set_Strings:
        i.set('SELECCIONA')
    for j in Set_Ints:
        j.set(0)
    Cantidad.set(1) 
    Total.set(0)
    Peso_Pescado.delete(0,END)
    Lectura = Comanda.get_children()
    Comanda.delete(*Lectura)
    Pagacon_Entry.delete(0,END)
    Tickets_box.delete(0,END)

def Lock_Unlock_Workspace(event):
    if event == 1:  ##LOCK WORKSPACE
        Boton_Agregar.config(state = 'disable')
        Cancel_Boton.configure(state = 'normal')
        Save_Ticket_bot.config(state = 'normal')
        Pagoefect.config(state = 'normal')
        Pagotar.config(state = 'normal')
        Meseros_Lista.config(state = 'disable')
        Env3Boton1.config(state = 'disable')
        Env3Boton2.config(state = 'disable')
        
    else:
        Boton_Agregar.config(state = 'normal')
        Pagoefect.config(state = 'disable')
        Pagotar.config(state = 'disable')
        Meseros_Lista.config(state = 'normal')
        Save_Ticket_bot.config(state = 'disable')
        Cancel_Boton.configure(state = 'disable')
        Env3Boton1.config(state = 'normal')
        Env3Boton2.config(state = 'normal')
        Factura.set(0)
        MetPago.set(0)

def Check_metpago(*args):
    if MetPago.get() == 0:
        Pagacon_Entry.config(state = 'normal')
        
    else:
        Pagacon_Entry.delete(0,END)
        Pagacon_Entry.config(state = 'disable')
        Factura.set(TRUE)

def Generar_Ticket():

    Lec = Comanda.get_children()
    a = Total.get()
    state = 1
    Fecha = time.strftime("%d/%m/%Y, Hora:%H:%M")
    if Llevar.get() == 0 or Lec == () or Mesero.get() == 'SELECCIONA':
        if Lec == () or Personas.get() == 0 or Mesa.get() == 0 or Mesero.get() == 'SELECCIONA':
            messagebox.showerror('ERROR','Revise:\n *Responsable de pedido\n*Número de Mesa\n*Número de Personas\n*Productos a cobrar')
            state = 0
    if state == 1:
        Cobrar_Boton.tkraise()
        Lock_Unlock_Workspace(1)
        Generar_Registro(a)

        Process_1 = []
        Process_2 = []
        words = ['CANTIDAD','DESCRIPCION','PRECIO']
        
        for i in Lec:#####FORMAR UNA LISTA CON LAS LISTAS DEL CONTENIDO DE LA COMANDA
            Process_1.append(Comanda.item(i)["values"])

        
        for x in Process_1:
            ZipObj = zip(words, x)
            dicts = dict(ZipObj)
            Process_2.append(dicts)
        
        #Actualización
        data = {"Orden": Process_2,"Info":[Fecha,Mesa.get(),Personas.get(),Descuento.get(),Importe.get()]}
        Process_1.clear() #### SE LIMPIA LA LISTA DEL PROCESO 1

        try:
            url = 'http://localhost/thermal_printer/Printer.php'
            requests.post(url,params=urlencode(data))
        except:
            messagebox.showerror(title='ERROR', message='NO SE PUDO IMPRIMIR EL TICKET')
        Process_2.clear()### LIMPIEZA DE PROCESS 2

def Generar_Registro(tot):
    Fecha = time.strftime("%d/%b/%Y, Hora:%H:%M:%S")
    Lec = Comanda.get_children()
    ############################################################ Generar Código de Venta
    cursor.execute("SELECT * FROM REGISTRO ORDER BY COD_VENTA DESC LIMIT 1")
    Last_ID = cursor.fetchone()[0]
    Last_number = int(Last_ID[8:12])

    if (Last_ID[0:8] == time.strftime("%Y%m%d")): 
        Last_number += 1
    else:
        Last_number =  1
    
    New_ID = time.strftime("%Y%m%d") + str(Last_number).zfill(3)
    
    ############################################################

    
    #cursor.execute("SELECT MAX(COD_VENTA) FROM REGISTRO")
    #New_ID = cursor.fetchone()[0] + 1


    Tipos = ('MESAS','LLEVAR')
    Tipo = Tipos[Llevar.get()]
    No_mesa = Mesa.get()
    Id_Mesero = cursor.execute("SELECT id FROM EMPLEADOS WHERE NOMBRE = ?",[Mesero.get()]).fetchone()[0]
    Ganancia = tot
    Imp = Importe.get()
    Des = Descuento.get()
    Pagos = ('EFECTIVO','TARJETA')
    Met_Pago = Pagos[MetPago.get()]
    No_Per = Personas.get()
    Fact_OPS = ('NO','SI')
    Fact = Fact_OPS[Factura.get()]

    items = []
    subitems = []
    superitems = []
    
    cursor.execute("INSERT INTO REGISTRO (COD_VENTA, ID_MESA, TIPO, ID_EMPLEADO, GANANCIA,IMPORTE, DESCUENTO, NUM_PERSONAS, MET_PAGO,FACTURA ,FECHA) VALUES(?,?,?,?,?,?,?,?,?,?,?)",(New_ID,No_mesa,Tipo,Id_Mesero,Ganancia,Imp,Des,No_Per,Met_Pago,Fact,Fecha))
    db.commit()
    count = 0
    for i in Lec:
        subitems.append(Comanda.item(i)['values'])
        items.append(New_ID)
        items.append(No_mesa)
        for j in subitems[count]: 
            items.append(j)
        superitems.append(items)
        count += 1
        items = []
    cursor.executemany("INSERT INTO MESA_CONSUMO (COD_VENTA, ID_MESA,CANTIDAD,PRODUCTO,COSTO) VALUES(?,?,?,?,?)",superitems)
    db.commit()    

def Save_Workspace():
    Lock_Unlock_Workspace(2)
    Generar_Boton.tkraise()
    cursor.execute("SELECT MAX(ID_Ticket) FROM TICKETS") #FORMA UN NUEVO ID PARA TICKETS
    New_Ticket_ID = cursor.fetchone()[0] + 1
    cursor.execute("SELECT MAX(REGISTRO.COD_VENTA) FROM REGISTRO") #IDENTIFICA EN QUÉ COD_VENTA ESTAMOS TRABAJANDO
    Current_Id_venta = cursor.fetchone()[0]

    tot = Total.get()
    Mes = cursor.execute("SELECT id FROM EMPLEADOS WHERE NOMBRE = ?",[Mesero.get()]).fetchone()[0]

    if AddDiscount.get() == FALSE:
        Des_op = 0
    else:
        Des_op = 1
    Des = Descuento.get()
    if AddAmount.get() == False:
        Imp_op = 0
    else:
        Imp_op = 1
    Imp = Importe.get()
    Per = Personas.get()
    No_mesa = Mesa.get()
    if Llevar.get() == FALSE:
        Llevar_op = 0
    else:
        Llevar_op = 1
        
    Saving = [No_mesa, New_Ticket_ID,Current_Id_venta, tot, Mes, Des_op, Des, Imp_op, Imp, Per, Llevar_op]
    
    cursor.execute("INSERT INTO TICKETS (NO_MESA, ID_Ticket, COD_VENTA, GANANCIA, EMPLEADO_ID, DES_OP, DESCUENTO, IMP_OP, IMPORTE, NO_PER, LLEVAR_OP) VALUES(?,?,?,?,?,?,?,?,?,?,?)",Saving)
    db.commit()

    UpdateTickets()
    Clear_WorkSpace()    

def Reload_Workspace():
    Id_Mesa = 0
    C_venta = 0
    Gan = 0
    Em_Id = 0
    Des_op = 0
    Des = 0
    Imp_op = 0
    Imp = 0
    Client_op = 0
    Client = ""
    Per = 0
    Llevar_op = 0
    Lec = Comanda.get_children()

    if Lec == ():
        try:
            CurrentTicket_ID = ticket_temp.get()
            Current_COD_venta = cursor.execute("SELECT COD_VENTA FROM TICKETS WHERE ID_Ticket = ?",[CurrentTicket_ID]).fetchone()[0]

            if not CurrentTicket_ID == 0 or CurrentTicket_ID == "":
                
                Box = [Id_Mesa, CurrentTicket_ID,C_venta, Gan, Em_Id, Des_op, Des, Imp_op, Imp, Client_op, Client,Per,Llevar_op]    
                count = 0
                cursor.execute("SELECT * FROM TICKETS WHERE ID_Ticket = ?",[CurrentTicket_ID])
                Data = cursor.fetchone()
                for x in Data:
                    Box[count] = x
                    count += 1

                cursor.execute("SELECT CANTIDAD, PRODUCTO, COSTO FROM MESA_CONSUMO WHERE COD_VENTA = ?",[Current_COD_venta])
                for i in cursor.fetchall():
                    Comanda.insert("",END, values = i)

                Mesa.set(Box[0])
                Total.set(Box[3])
                cursor.execute("SELECT NOMBRE FROM EMPLEADOS WHERE id = ?",[Box[4]])
                Mesero.set(cursor.fetchone()[0])
                AddDiscount.set(Box[5])
                Descuento.set(Box[6])
                AddAmount.set(Box[7])
                Importe.set(Box[8])
                    
                Lock_Unlock_Workspace(1)
                Cobrar_Boton.tkraise()
            else:
                messagebox.showerror(title='ERROR', message='DEBES SELECCIONAR UN NÚMERO DE TICKET')
        except TclError:
            messagebox.showerror(title='ERROR', message='DEBES SELECCIONAR UN NÚMERO DE TICKET')
    else:
        messagebox.showerror(title='ERROR', message='CAMPO OCUPADO, POR FAVOR, PRIMERO COBRA ESTE TICKET')

def Cobrar_ticket():
    Generar_Boton.tkraise()
    try:
        if float(Pagacon_Entry.get()) != 0:
            Devolver = "$" + str(float(Pagacon_Entry.get()) - Total.get())
            
        else:
            Devolver = 0
    except ValueError:
        Devolver = 0
        pass
    try:
        Ticket_ID = ticket_temp.get()
        if not Ticket_ID == 0:
            Cod_ID = cursor.execute("SELECT COD_VENTA FROM TICKETS WHERE ID_Ticket = ?", [Ticket_ID]).fetchone()[0]
            LastaUpdate_Registro(MetPago.get(), Factura.get(), Cod_ID)
        else:
            Ticket_ID = cursor.execute("SELECT MAX(COD_VENTA) FROM REGISTRO").fetchone()[0]
            LastaUpdate_Registro(MetPago.get(), Factura.get(), Ticket_ID)
    except TclError:
        Ticket_ID = cursor.execute("SELECT MAX(COD_VENTA) FROM REGISTRO").fetchone()[0]
        LastaUpdate_Registro(MetPago.get(), Factura.get(), Ticket_ID)

    
    Borrar_ticket(Ticket_ID)
    Clear_WorkSpace()
    Lock_Unlock_Workspace(2)
    def Closewindow():
        Cobro_window.destroy()

    Cobro_window = Toplevel()
    Cobro_window.geometry('300x160')
    Cobro_window.title('COBRAR...')
    Cobro_window.config(bg = '#FEF5E7')
    Cobro_window.resizable(1,1)

    Mensajelbl = Label(Cobro_window, text = 'DEVOLVER LA CANTIDAD DE')
    Mensajelbl.config(bg = '#FEF5E7',font = 'Arial 10 bold')
    Mensajelbl.pack(pady = 10)

    Cantidadlbl = Label(Cobro_window, text = Devolver)
    Cantidadlbl.config(bg = '#FEF5E7',font = 'Arial 30 bold')
    Cantidadlbl.pack(pady = 10)

    okboton = Button(Cobro_window, text = '  OK  ', command = Closewindow)
    okboton.config(bg = '#FEF5E7',font = 'Arial 15 bold')
    okboton.pack(pady = 10,ipady=25)

    Cobro_window.protocol("WM_DELETE_WINDOW", Closewindow)

def UpdateTickets():
    del Tickets[:]
    temporal = []

    Codigos_venta = cursor.execute("SELECT COD_VENTA FROM TICKETS")
    for i in cursor.fetchall():
        temporal.append(i[0])
    
    for x in temporal:
        cursor.execute("SELECT ID_Ticket FROM TICKETS WHERE COD_VENTA = ?",[x])
        for j in cursor.fetchall():
            Tickets.append(j[0])
    
    Tickets_box['values'] = Tickets

def LastaUpdate_Registro(Met_Pago, a, Id):
    Pagos = ('EFECTIVO','TARJETA')
    Fact_ops = ('NO','SI')
    Factu = Fact_ops[a]
    Pago = Pagos[Met_Pago]
    cursor.execute("UPDATE REGISTRO SET MET_PAGO = ?, FACTURA = ? WHERE COD_VENTA = ?",(Pago, Factu, Id))
    db.commit()

def Cancel():

    Clear_WorkSpace()
    Lock_Unlock_Workspace(2)
    Generar_Boton.tkraise()

def Borrar_ticket(TickID):
    
    try:
        codID = cursor.execute("SELECT COD_VENTA FROM TICKETS WHERE ID_Ticket = ?",[TickID]).fetchone()[0]
        cursor.execute("DELETE FROM TICKETS WHERE COD_VENTA = ?",[codID])
        db.commit()
    except TypeError:
        #messagebox.showerror(title='ERROR', message='NO EXISTE CONTENIDO EN ESTA CASILLA')
        pass


###################################################################################
window = Tk() #INICIALIAR VENTANA
wx = window.winfo_screenwidth() 
wy= window.winfo_screenheight() 


Dx = 740
Dy = 700

#IMAGENES
Logo = PhotoImage(file = "Images/Logo.png")
DbLogo = PhotoImage(file = 'Images/DBLogo.png')
ExcelLogo = PhotoImage(file = 'Images/Excel.png')
#user_img = PhotoImage(file = 'Images/User.png')
#Key_img = PhotoImage(file = 'Images/Key.png')


########################## VENTANA PRINCIPAL######################
window.geometry('%ix%i'%(wx * 0.4,wy * 0.85))
window.title('Tickets Chalo Manager V1.4')
window.config(bg='#E9967A')
#window.resizable(False,False)

#ETIQUETAS
Title = Label(window,text='GENERADOR DE TICKETS CHALO')
Title.place(relx = 0.2,rely=0.01)
Title.config(width = '42', height = '2',bd=10, relief='sunken',bg = "#F08080", font="Arial 15 bold")

#Reloj_Eti = Label(window)
#Reloj_Eti.place(relx = 0.337, rely = 0.085)
#Reloj_Eti.config(bd=5, relief='sunken', bg='#E9967A', fg="white")
#Display_Time()

LogoLabel = Label(image=Logo)
LogoLabel.place(relx= (330/Dx) ,rely=(95/Dy))
LogoLabel.config(bd=2,relief='solid')

Etiqueta2 = Label(window,text= 'PAGA CON: $')
Etiqueta2.config(bg='#E9967A',font='Arial 11 bold')
Etiqueta2.place(relx=485/Dx,rely=155/Dy)

Etiqueta3 = Label(window, text= 'Registros')
Etiqueta3.config(bg='#E9967A',font='Arial 8 bold')
Etiqueta3.place(relx=545/Dx,rely=75/Dy)

#Etiqueta4 = Label(window, text = 'Conectar')
#Etiqueta4.config(bg='#E9967A',font='Arial 12 bold')
#Etiqueta4.place(relx = 645/Dx,rely=75/Dy)

Etiqueta5 = Label(window, text = 'TOTAL')
Etiqueta5.config(bg='#E9967A',font='Arial 10 bold')
Etiqueta5.place(relx = 355/Dx, rely = 220/Dy)

Total = DoubleVar()
Total.set(0)
Total.trace('w', Connect_Tot)
Output_Frame = Label(window, textvariable = Total)
Output_Frame.config(relief='sunken', bg='#FEF5E7', font = 'Arial 12 bold')
Output_Frame.place(relx = 350/Dx, rely = 250/Dy, relwidth = 0.1, relheight = 0.04)
Peso_Marca = Label(window, text = '$')
Peso_Marca.config(bg='#E9967A', font = 'Arial 10 bold')
Peso_Marca.place(relx = 330/Dx, rely = 250/Dy)

#ENTRADAS

Pagacon_Entry = Entry(window, text = 0)
Pagacon_Entry.config(font='Arial 15 bold',justify="center",fg='green')
Pagacon_Entry.place(relx=620/Dx,rely=148/Dy, relwidth=0.1)

#VERIFICACIÓN
MetPago = IntVar()
Pagoefect = Radiobutton(window, text = 'Efec.',variable = MetPago, value = 0)
Pagoefect.config(bg = '#E9967A', state = 'disable')
Pagoefect.place(relx=484/Dx,rely=200/Dy)

Pagotar = Radiobutton(window, text = 'Tarjeta', variable = MetPago, value = 1)
Pagotar.config(bg = '#E9967A',state = 'disable')
Pagotar.place(relx = 560/Dx, rely = 200/Dy)
MetPago.trace('w', Check_metpago)

Factura = BooleanVar()
Factura.set(FALSE)
Fact_ver = Checkbutton(window, text = 'Factura', var = Factura)
Fact_ver.config(bg = '#E9967A')
Fact_ver.place(relx = 640/Dx, rely = 200/Dy)

#BOTONES
Generar_Boton = Button(window, text = "Generar Ticket", command = Generar_Ticket)
Generar_Boton.config(bg='#b2f2bb',font='Arial 10 bold')
Generar_Boton.place(relx=180/Dx,rely=660/Dy, width = 120, height = 25 )

Cobrar_Boton = Button(window, text = '  Cobrar  ', command = Cobrar_ticket)
Cobrar_Boton.config(bg='#E9967A',font='Arial 10 bold',fg = 'Green')
Cobrar_Boton.place(relx=180/Dx,rely=660/Dy, width = 120, height = 25 )

Generar_Boton.tkraise()

Cancel_Boton = Button(window, text = 'Cancelar', command = Cancel)
Cancel_Boton.config(bg = '#E9967A',font='Arial 9 bold',fg = 'red', state = 'disable')
Cancel_Boton.place(relx = 340/Dx, rely = 660/Dy, width = 70, height = 25)

AdminDbBoton = Button(window, command = Data_Window)
AdminDbBoton.config(image = DbLogo)
AdminDbBoton.place(relx=555/Dx, rely = 93/Dy)

ExcelConnect = Button(window, command=ConnectExcel)
ExcelConnect.config(image = ExcelLogo) 
ExcelConnect.place(relx=660/Dx,rely= 93/Dy)

Save_Ticket_bot = Button(window, text = "Guardar Ticket", command = Save_Workspace)
Save_Ticket_bot.config(fg = 'green', state='disabled')
Save_Ticket_bot.place(relx = 460/Dx, rely = 232/Dy, width = 100, height = 20)

Reload_bot = Button(window, text = 'Re', command = Reload_Workspace)
Reload_bot.config(fg = 'blue', state = 'normal')
Reload_bot.place(relx = 690/Dx, rely = 232/Dy, width = 30, height = 20)

#LISTA
ticket_temp = IntVar()
ticket_temp.set(0)
Tickets_box = ttk.Combobox(window,textvariable = ticket_temp, values = Tickets, postcommand = UpdateTickets)
Tickets_box.config(justify = 'center',font = 'Arial 11 ')
Tickets_box.place(relx = 585/Dx, rely = 232/Dy, width = 80, height = 20)

#------------------------------------ENVOLTURA 1-------------------------------#
Envoltura1 = LabelFrame(window,text='Detalles de la Orden')
Envoltura1.config(bg='#E9967A')
Envoltura1.place(relx=5/Dx,rely=90/Dy, relwidth=0.412,relheight=0.26)

Dim_E1y = 182 # DIMENSIONES DE LA ENVOLTURA 1
Dim_E1x = 304.88

#ETIQUETAS
Env1Eti1 = Label(Envoltura1,text = 'Responsable:') 
Env1Eti1.config(bg='#E9967A',font='Arial 8 bold')
Env1Eti1.place(relx=12/Dim_E1x,rely=10/Dim_E1y)

Env1Eti2 = Label(Envoltura1, text = 'No.Mesa')
Env1Eti2.config(bg='#E9967A',font='Arial 8 bold')
Env1Eti2.place(relx=230/Dim_E1x,rely=10/Dim_E1y)

Env1Eti3 = Label(Envoltura1, text = 'No.Per.')
Env1Eti3.config(bg = '#E9967A', font = 'Arial 8 bold')
Env1Eti3.place(relx = 145/Dim_E1x, rely = 10/Dim_E1y)

#VERIFICACIÓN
AddDiscount = BooleanVar()
AddDiscount.set(FALSE)
Env1Ver1 = Checkbutton(Envoltura1,text="Descuento($pesos):",var=AddDiscount, command = CheckDescuento)
Env1Ver1.config(bg='#E9967A',font='Arial 8 bold')
Env1Ver1.place(relx=5/Dim_E1x,rely=70/Dim_E1y)

AddAmount = BooleanVar()
AddAmount.set(FALSE)
Env1Ver2 = Checkbutton(Envoltura1,text="Importe($pesos):",var = AddAmount, command = CheckImporte)
Env1Ver2.config(bg='#E9967A',font='Arial 8 bold')
Env1Ver2.place(relx=5/Dim_E1x,rely=100/Dim_E1y)

Llevar = BooleanVar()
Llevar.set(FALSE)
Env1Ver4 = Checkbutton(Envoltura1, text = 'Llevar', var = Llevar, command = CheckLlevar)
Env1Ver4.config(bg='#E9967A',font='Arial 8 bold')
Env1Ver4.place(relx = 220/Dim_E1x, rely = 85/Dim_E1y)
#ENTRADAS DE TEXTO
Descuento = DoubleVar()
Descuento.set(0)
Env1Ent1 = Entry(Envoltura1, textvariable = Descuento)
Env1Ent1.config(justify="center", font="Arial 15", state = 'disable')
Env1Ent1.place(relx = 170/Dim_E1x, rely = 65/Dim_E1y,relwidth = 0.15, relheight=0.17)
Descuento.trace('w',Connect_Tot)
Importe = DoubleVar()
Importe.set(0)
Env1Ent2 = Entry(Envoltura1, textvariable = Importe)
Env1Ent2.config(justify="center", font="Arial 15", state = 'disable')
Env1Ent2.place(relx = 170/Dim_E1x, rely = 98/Dim_E1y,relwidth = 0.15, relheight=0.17)
Importe.trace('w',Connect_Tot)
Personas = IntVar()
Personas.set(0)
Ent_Personas = Entry(Envoltura1, textvariable = Personas)
Ent_Personas.config(justify = 'center', font = 'Arial 11 bold')
Ent_Personas.place(relx = 155/Dim_E1x, rely = 30/Dim_E1y, relwidth = 0.1, relheight = 0.15)

Mesa = IntVar()
Mesa.set(0)
Mesas_Ent = Entry(Envoltura1,textvariable = Mesa)
Mesas_Ent.config(justify = 'center',font='Arial 11 bold')
Mesas_Ent.place(relx=250/Dim_E1x,rely=30/Dim_E1y,relwidth=0.1, relheight = 0.15)

#LISTAS
Mesero = StringVar()
Mesero.set('SELECCIONA')
Meseros_Lista = ttk.Combobox(Envoltura1, textvariable = Mesero, values = Mesero, postcommand = UpdateListas)
Meseros_Lista.config(justify = 'center',font = 'Arial 8 ')
Meseros_Lista.place(relx=5/Dim_E1x,rely=30/Dim_E1y,relwidth=0.4, relheight = 0.15)


#BOTONES
MasPerBoton = Button(Envoltura1, text = '+', command = lambda: SisPersonas(1))
MasPerBoton.config(fg = 'blue')
MasPerBoton.place(relx = 185/Dim_E1x, rely = 30/Dim_E1y, relwidth = 0.06, relheight = 0.15)

MenPerBoton = Button(Envoltura1, text = '-', command = lambda: SisPersonas(2))
MenPerBoton.config(fg = 'red')
MenPerBoton.place(relx = 139/Dim_E1x, rely = 30/Dim_E1y, relwidth = 0.06, relheight = 0.15)

MasMesBoton = Button(Envoltura1, text = '>', command = lambda: SisMesas(1))
MasMesBoton.config(fg = "blue")
MasMesBoton.place(relx = 276/Dim_E1x, rely = 30/Dim_E1y, relwidth = 0.06, relheight = 0.15)

MenMesBoton = Button(Envoltura1, text = '<', command = lambda: SisMesas(2))
MenMesBoton.config(fg = "blue")
MenMesBoton.place(relx = 235/Dim_E1x, rely = 30/Dim_E1y, relwidth = 0.06, relheight = 0.15)
#------------------------------------ENVOLTURA 2-------------------------------#

Envoltura2 = LabelFrame(window, text = 'Tomar Orden')
Envoltura2.config(bg='#E9967A')
Envoltura2.place(relx=5/Dx,rely=275/Dy,relwidth=0.6,relheight=0.54)

Dim_E2x = 310.8
Dim_E2y = 370.44

Contendor_Productos = Frame(Envoltura2)
Contendor_Productos.config(bg='#E9967A',relief=RAISED,bd=3)
Contendor_Productos.place(relx=20/444,rely=3/378,relwidth=0.7,relheight=0.98)

#ETIQUETAS
Env2Eti1 = Label(Contendor_Productos,text = "Cocteles")
Env2Eti1.config(bg='#E9967A',font='Arial 8 bold')
Env2Eti1.place(relx=120/Dim_E2x,rely=5/Dim_E2y)

Env2Eti2 = Label(Contendor_Productos,text = "Bebidas")
Env2Eti2.config(bg='#E9967A',font='Arial 8 bold')
Env2Eti2.place(relx=120/Dim_E2x,rely=46/Dim_E2y)

Env2Eti3 = Label(Contendor_Productos,text="Platillos")
Env2Eti3.config(bg='#E9967A',font='Arial 8 bold')
Env2Eti3.place(relx=120/Dim_E2x,rely=90/Dim_E2y)

Env2Eti4 = Label(Contendor_Productos,text="Caldos")
Env2Eti4.config(bg='#E9967A',font='Arial 8 bold')
Env2Eti4.place(relx=120/Dim_E2x,rely=134/Dim_E2y)

Env2Eti5 = Label(Contendor_Productos,text = 'Pescados')
Env2Eti5.config(bg='#E9967A',font='Arial 8 bold')
Env2Eti5.place(relx=60/Dim_E2x,rely=218/Dim_E2y)

Env2Eti6 = Label(Contendor_Productos,text ='Postres')
Env2Eti6.config(bg='#E9967A',font='Arial 8 bold')
Env2Eti6.place(relx=120/Dim_E2x,rely=175/Dim_E2y)

Env2Eti7 = Label(Contendor_Productos, text = 'Peso:              gr')
Env2Eti7.config(bg='#E9967A',font='Arial 8 bold')
Env2Eti7.place(relx=170/Dim_E2x,rely=240/Dim_E2y)

Env2Eti8 = Label(Contendor_Productos,text='Nombre')
Env2Eti8.config(bg= '#E9967A',font='Arial 8 bold')
Env2Eti8.place(relx=60/Dim_E2x,rely=293/Dim_E2y)

Env2Eti9 = Label(Contendor_Productos,text='Precio')
Env2Eti9.config(bg= '#E9967A',font='Arial 8 bold')
Env2Eti9.place(relx=220/Dim_E2x,rely=293/Dim_E2y)

Env2Eti10 = Label(Envoltura2, text = 'Cantidad')
Env2Eti10.config(bg = '#E9967A', font = 'Arial 10 bold')
Env2Eti10.place(relx = 343/444, rely = 90/378)

#LISTAS

Coctel = StringVar()
Coctel.set(Cocteles[0])
Cocteles_Lista = ttk.Combobox(Contendor_Productos, textvariable = Coctel, values = Cocteles, postcommand = UpdateListas)
Cocteles_Lista.config(font='Arial 8',justify="center")
Cocteles_Lista.place(relx=15/Dim_E2x, rely=23/Dim_E2y,relwidth = 0.9)

Bebida = StringVar()
Bebida.set(Bebidas[0])
Bebidas_Lista = ttk.Combobox(Contendor_Productos,textvariable=Bebida,values=Bebidas,postcommand = UpdateListas)
Bebidas_Lista.config(font='Arial 8',justify="center")
Bebidas_Lista.place(relx=15/Dim_E2x,rely=65/Dim_E2y,relwidth=0.9)

Platillo = StringVar()
Platillo.set(Platillos[0])
Platillos_Lista = ttk.Combobox(Contendor_Productos,textvariable = Platillo,values = Platillos,postcommand = UpdateListas)
Platillos_Lista.config(font='Arial 8',justify="center")
Platillos_Lista.place(relx=15/Dim_E2x,rely=108/Dim_E2y,relwidth=0.9)

Caldo = StringVar()
Caldo.set(Caldos[0])
Caldos_Lista = ttk.Combobox(Contendor_Productos,textvariable = Caldo,values = Caldos,postcommand = UpdateListas)
Caldos_Lista.config(font='Arial 8',justify="center")
Caldos_Lista.place(relx=15/Dim_E2x,rely=151/Dim_E2y,relwidth=0.9)

Pescado = StringVar()
Pescado.set(Pescados[0])
Pescado.trace('w',Activar_Peso)
Pescados_Lista = ttk.Combobox(Contendor_Productos,textvariable = Pescado, values = Pescados,postcommand = UpdateListas)
Pescados_Lista.config(font='Arial 8',justify="center")
Pescados_Lista.place(relx=15/Dim_E2x,rely=237/Dim_E2y,relwidth=0.5)

Postre = StringVar()
Postre.set(Postres[0])
Postres_Lista = ttk.Combobox(Contendor_Productos, textvariable = Postre, values = Postres,postcommand = UpdateListas)
Postres_Lista.config(font='Arial 8',justify="center")
Postres_Lista.place(relx=15/Dim_E2x,rely=194/Dim_E2y,relwidth=0.9)

#ENTRADAS
Peso_Pescado = Entry(Contendor_Productos, state = 'disable')
Peso_Pescado.config(justify='center')
Peso_Pescado.place(relx=220/Dim_E2x,rely=237/Dim_E2y,relwidth=0.15)

Nombre_Otro = Entry(Contendor_Productos, state = 'disable')
Nombre_Otro.config(justify = 'center')
Nombre_Otro.place(relx = 15/Dim_E2x, rely = 310/Dim_E2y, relheight= 0.08)

Precio_Otro = Entry(Contendor_Productos, state = 'disable')
Precio_Otro.config(justify = 'center')
Precio_Otro.place(relx=220/Dim_E2x,rely=310/Dim_E2y, relheight = 0.08, relwidth = 0.2)

Cantidad = IntVar()
Cantidad.set(1)
Cantidad_Entrada = Entry(Envoltura2, textvariable = Cantidad)
Cantidad_Entrada.config(justify = 'center', font = 'Arial 13 bold')
Cantidad_Entrada.place(relx = 360/444, rely = 119/378, relwidth = 0.12, relheight = 0.08)

#BOTONES
Boton_Agregar = Button(Envoltura2,text='Agregar', command = Agregar)
Boton_Agregar.config(fg='green',font='Arial 10 bold')
Boton_Agregar.place(relx=342/444,rely=150/378,relwidth=0.2)

CantidadmasBoton = Button(Envoltura2,text = '>', command = lambda: SisCantidades(1))
CantidadmasBoton.config(fg='blue',font='Arial 10 bold')
CantidadmasBoton.place(relx=415/444,rely=120/378, relwidth = 0.035)

CantidadmenosBoton = Button(Envoltura2,text = '<', command = lambda: SisCantidades(2))
CantidadmenosBoton.config(fg='blue',font='Arial 10 bold')
CantidadmenosBoton.place(relx=343/444,rely=120/378, relwidth = 0.035)

#VERIFICACIÓN
AddOther = BooleanVar()
AddOther.set(FALSE)
Env2Ver1 = Checkbutton(Contendor_Productos, text= 'Agregar Otro producto', var = AddOther)
Env2Ver1.config(bg='#E9967A',font='Arial 8 bold', command = Otro_Producto)
Env2Ver1.place(relx=20/Dim_E2x,rely=270/Dim_E2y)


#------------------------------------ENVOLTURA 3-------------------------------#

Envoltura3 = LabelFrame(window, text = 'Comanda')
Envoltura3.config(bg='#E9967A')
Envoltura3.place(relx=450/Dx,rely=255/Dy,relwidth=0.38,relheight=0.6)

Dim_E3x = 281.2
Dim_E3y = 420

#Lista
Comanda = ttk.Treeview(Envoltura3,columns=(1,2,3),show='headings')
Comanda.place(relx=1/Dim_E3x,rely=0,relwidth=0.98,relheight=0.9)
Comanda.column(1,width=25)
Comanda.column(2,width=100)
Comanda.column(3,width =30)
Comanda.heading(1, text="Cant.")
Comanda.heading(2, text="Producto")
Comanda.heading(3, text='Precio')

#BOTONES
Env3Boton1 = Button(Envoltura3,text = 'Borrar', command = lambda: Borrar(1))
Env3Boton1.config(fg='red')
Env3Boton1.place(relx=200/Dim_E3x,rely=380/Dim_E3y)

Env3Boton2 = Button(Envoltura3,text= 'Limpiar Comanda', command = lambda: Borrar(2))
Env3Boton2.config(fg='red')
Env3Boton2.place(relx=20/Dim_E3x,rely=380/Dim_E3y)


window.mainloop()
